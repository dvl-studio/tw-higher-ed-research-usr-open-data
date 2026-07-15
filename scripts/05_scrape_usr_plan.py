from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import requests
import urllib3
from openpyxl import Workbook


API_URL = "https://usr.moe.gov.tw/API/Project/tw"
BASE_URL = "https://usr.moe.gov.tw"
OUTPUT_FILE = "USR plan.xlsx"
TARGET_PHASE_TITLES = {"第二期", "第三期", "第四期"}
PHASE_ORDER = {"第二期": 2, "第三期": 3, "第四期": 4}

VERIFY_SSL = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def normalize_phase_title(title: str) -> str:
    return title.replace("第", "").replace("期", "").strip()


def request_json(method: str = "get", **kwargs: Any) -> dict[str, Any]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
        ),
        "Referer": "https://usr.moe.gov.tw/tw/plan/maps?vid=4",
    }
    response = requests.request(method, API_URL, headers=headers, timeout=60, verify=VERIFY_SSL,**kwargs)
    response.raise_for_status()
    payload = response.json()
    if not payload.get("status"):
        raise RuntimeError(f"API returned unsuccessful status: {payload!r}")
    return payload["data"]


def get_target_phases() -> list[dict[str, Any]]:
    data = request_json()
    phases = data.get("Plane", [])
    targets = []
    for phase in phases:
        title = str(phase.get("plane_title", "")).strip()
        normalized = normalize_phase_title(title)
        if title in TARGET_PHASE_TITLES or normalized in {"二", "三", "四"}:
            targets.append(phase)
    return sorted(targets, key=lambda item: int(item["id"]))


def iter_report_dicts(report_data: Any):
    if isinstance(report_data, dict):
        report_groups = report_data.values()
    elif isinstance(report_data, list):
        report_groups = report_data
    else:
        return

    for group in report_groups:
        if not isinstance(group, dict):
            continue
        content = group.get("content", [])
        if isinstance(content, dict):
            yield from content.values()
        elif isinstance(content, list):
            yield from content


def build_doc_url(doc_file: Any) -> str:
    doc_path = str(doc_file or "").strip()
    if not doc_path:
        return ""
    if doc_path.startswith(("http://", "https://")):
        return doc_path
    if doc_path.startswith("/"):
        return f"{BASE_URL}{doc_path}"
    return f"{BASE_URL}/{doc_path}"


def fetch_phase_rows(phase: dict[str, Any]) -> list[dict[str, str]]:
    phase_id = str(phase["id"])
    phase_title = str(phase["plane_title"]).strip()
    data = request_json("post", data={"plane": phase_id})

    rows = []
    seen_report_ids = set()
    for report in iter_report_dicts(data.get("Report")):
        if not isinstance(report, dict):
            continue

        report_id = report.get("id")
        if report_id is None:
            report_id = (
                str(report.get("year", "")).strip(),
                str(report.get("school", "")).strip(),
                str(report.get("title", "")).strip(),
                str(report.get("doc_file", "")).strip(),
            )
        if report_id in seen_report_ids:
            continue
        seen_report_ids.add(report_id)

        rows.append(
            {
                "期別": phase_title,
                "年份": str(report.get("year", "")).strip(),
                "學校名稱": str(report.get("school", "")).strip(),
                "報告連結": build_doc_url(report.get("doc_file")),
            }
        )
    return rows


def write_excel(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "USR plan"

    headers = ["期別", "年份", "學校名稱", "報告連結"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])
        link_cell = sheet.cell(row=sheet.max_row, column=4)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 12, "B": 10, "C": 32, "D": 90}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for cell in sheet[1]:
        cell.style = "Headline 3"

    workbook.save(output_path)


def main() -> int:
    phases = get_target_phases()
    if not phases:
        raise RuntimeError("Cannot find target phases: 第二期, 第三期, 第四期")

    rows = []
    for phase in phases:
        phase_rows = fetch_phase_rows(phase)
        rows.extend(phase_rows)
        print(f"{phase['plane_title']}: {len(phase_rows)} rows")

    rows.sort(
        key=lambda row: (
            PHASE_ORDER.get(row["期別"], 99),
            row["年份"],
            row["學校名稱"],
        )
    )

    project_root = Path(__file__).resolve().parents[1]
    output_path = project_root / "data" / "raw" / "05_USR_plan.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(rows, output_path)
    print(f"Saved {len(rows)} rows to {output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
